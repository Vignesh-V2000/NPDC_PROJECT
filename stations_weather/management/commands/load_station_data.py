import os
import re
import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

from django.core.management.base import BaseCommand
from django.db import transaction
from django.utils import timezone
from pytz import timezone as pytz_timezone

# Import the existing Django models
from stations_weather.models import Last24HrsData, BharatiWeatherData

IST = pytz_timezone('Asia/Kolkata')


def make_safe_combine(date_obj):
    def safe_combine(time_val):
        try:
            if pd.isna(time_val) or str(time_val).strip() == "":
                return pd.NaT
            time_part = pd.to_datetime(str(time_val), errors='coerce')
            if pd.isna(time_part):
                return pd.NaT
            return datetime.combine(date_obj.date(), time_part.time())
        except Exception:
            return pd.NaT
    return safe_combine

def extract_date(filename):
    match = re.search(r'(\d{2}-[A-Za-z]{3}-\d{4})', filename)
    return datetime.strptime(match.group(1), "%d-%b-%Y") if match else None


class Command(BaseCommand):
    help = 'Process DCWIS data and load it into the local NPDC database (SQLite or PostgreSQL)'

    def add_arguments(self, parser):
        parser.add_argument("-r", "--rerun", action="store_true", help="Rerun for last 5 days data")
        parser.add_argument(
            '--mock-dir', 
            type=str, 
            default='raw_data/DCWIS_NEW', 
            help='Relative path to the directory containing raw Excel files'
        )

    def handle(self, *args, **options):
        # We assume you'll place the raw files in a folder like C:/.../NPDC_PROJECT/raw_data/DCWIS_NEW
        from django.conf import settings
        
        # Absolute paths for our local run
        DATA_DIR = os.path.join(settings.BASE_DIR, options['mock_dir'])
        TEMPLATE_FILE = os.path.join(DATA_DIR, "template.xlsx")
        OUTPUT_DIR = os.path.join(settings.BASE_DIR, "raw_data", "DCWIS")
        
        # Ensure directories exist
        os.makedirs(DATA_DIR, exist_ok=True)
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        
        # If the template file isn't there, we can't fully replicate the script without it
        if not os.path.exists(TEMPLATE_FILE):
             self.stdout.write(self.style.ERROR(f"Template file not found at {TEMPLATE_FILE}."))
             self.stdout.write(self.style.WARNING(f"Please copy the real 'template.xlsx' to {DATA_DIR} and rerun."))
             return
             
        # Check if the raw data directory has files
        rw09_files = {extract_date(f): os.path.join(DATA_DIR, f)
                    for f in os.listdir(DATA_DIR) if f.startswith("RW09_") and f.endswith(".xlsx")}
        wind_files = {extract_date(f): os.path.join(DATA_DIR, f)
                    for f in os.listdir(DATA_DIR) if f.startswith("Wind_") and f.endswith(".xlsx")}

        if not rw09_files or not wind_files:
            self.stdout.write(self.style.WARNING(f"No RW09 or Wind files found in {DATA_DIR}"))
            return

        common_dates = sorted(set(rw09_files) & set(wind_files), reverse=True)[:5]
        
        for date in common_dates:
            output_filename = date.strftime("%d-%b-%Y") + ".xlsx"
            output_path = os.path.join(OUTPUT_DIR, output_filename)

            if os.path.exists(output_path) and not options['rerun']:
                self.stdout.write(f"Skipping {output_filename} (already exists)")
                continue

            self.stdout.write(f"Processing {output_filename}")

            rw09_df = pd.read_excel(rw09_files[date])
            wind_df_raw = pd.read_excel(wind_files[date])

            wind_df = wind_df_raw.iloc[1:].copy()
            wind_df.columns = ['Date', 'Time', 'Wind Direction (DEG)', 'WD 2Min', 'WD 10Min',
                            'Wind Speed (knots)', 'WS 2Min', 'WS 10Min',
                            'WD at Max WS in 1 Min (DEG)', 'Max WS in 1Min (knots)']

            merged_df = pd.merge(rw09_df, wind_df, on=['Date', 'Time'], how='inner')
            formatted_date = date.strftime("Date : %d-%b-%Y")

            final_df = pd.DataFrame({
                "SR.No.": range(1, len(merged_df) + 1),
                "Time (HH:mm)": merged_df["Time"],
                "Wind Direction (Deg)": merged_df["Wind Direction (DEG)"],
                "Wind Speed (Knots)": merged_df["Wind Speed (knots)"],
                "Temperature (DegC)": merged_df["Temperature1MinAvg (DEG C)"],
                "Presssure (mbar)": merged_df["Pressure1MinAvg (mBar)"],
                "Humidity (%)": merged_df["Humidity1MinAvg (%Rh)"],
                "Dew Point (DegC)": merged_df["DewPoint1MinAvg (DEG C)"]
            })

            final_df.fillna({
                "Wind Direction (Deg)": "-",
                "Wind Speed (Knots)": "--.-",
                "Temperature (DegC)": "--.-",
                "Presssure (mbar)": "----.-",
                "Humidity (%)": "--.-",
                "Dew Point (DegC)": "--.-"
            }, inplace=True)

            wb = load_workbook(TEMPLATE_FILE)
            ws = wb.active

            for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), start=6):
                for c_idx, value in enumerate(row, start=2):
                    ws.cell(row=r_idx, column=c_idx, value=value)

            ws["B3"] = formatted_date
            wb.save(output_path)
            self.stdout.write(self.style.SUCCESS(f"Created merged file: {output_filename}"))
            
            # Sub-process: Load into DB
            self._load_into_db(output_filename, OUTPUT_DIR)


    def _load_into_db(self, file_name, directory_path):
        """Replaces the psycopg2 raw queries with Django ORM queries"""
        self.stdout.write(f"Loading {file_name} into Django database...")
        excel_path = os.path.join(directory_path, file_name)

        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            b3_value = ws["B3"].value

            if b3_value and "Date" in b3_value:
                date_part = b3_value.split(":")[-1].strip()
                date_obj = datetime.strptime(date_part, "%d-%b-%Y")
            else:
                raise ValueError("B3 empty or invalid")

        except Exception as e:
            file_stem = os.path.splitext(file_name)[0]
            date_obj = datetime.strptime(file_stem, "%d-%b-%Y")

        df = pd.read_excel(excel_path, skiprows=5)
        column_map = {
            "Time (HH:mm)": "obstime",
            "Temperature (DegC)": "tempr",
            "Presssure (mbar)": "ap",
            "Wind Speed (Knots)": "ws",
            "Wind Direction (Deg)": "wd",
            "Humidity (%)": "rh"
        }
        df = df.rename(columns=column_map)
        df = df[list(column_map.values())]

        for col in ['tempr', 'ap', 'ws', 'wd', 'rh']:
            df[col] = df[col].apply(lambda x: np.nan if pd.isna(x) else x)
            
        numeric_cols = ['tempr', 'ap', 'ws', 'wd', 'rh']
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        
        for col in numeric_cols:
            df[col] = df[col].apply(lambda x: x if pd.notna(x) else -999)
            
        df['obstime'] = df['obstime'].apply(make_safe_combine(date_obj))
        df['obstime'] = pd.to_datetime(df['obstime'], errors='coerce')
        df = df.where(pd.notnull(df), None)

        inserted = 0
        updated = 0

        # We use a transaction to ensure all inserts/updates for this file succeed together
        with transaction.atomic():
            for index, row in df.iterrows():
                if pd.isna(row['obstime']):
                    continue
                
                dt_aware = IST.localize(row['obstime'])
                
                # Last24HrsData Table mapping (uses TimeField and DateField separately)
                time_val = dt_aware.time()
                date_val = dt_aware.date()
                
                # imd_bharati Table mapping (uses full DateTimeField)
                obstime_full = dt_aware
                
                # Update or crate for imd_bharati
                obj, created = BharatiWeatherData.objects.update_or_create(
                    obstime=obstime_full,
                    defaults={
                        'tempr': row['tempr'],
                        'ap': row['ap'],
                        'ws': row['ws'],
                        'wd': row['wd'],
                        'rh': row['rh']
                    }
                )
                
                # The official script also tries to insert to Last24HrsData
                Last24HrsData.objects.update_or_create(
                    obstime=time_val,
                    date=date_val,
                    defaults={
                        'tempr': row['tempr'],
                        'ap': row['ap'],
                        'ws': row['ws'],
                        'wd': row['wd'],
                        'rh': row['rh']
                    }
                )

                if created:
                    inserted += 1
                else:
                    updated += 1
                    
        self.stdout.write(self.style.SUCCESS(f"Finished {file_name}: Inserted {inserted}, Updated {updated} records."))
