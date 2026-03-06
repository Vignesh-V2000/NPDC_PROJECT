# import schedule
import time
import sys
from pathlib import Path
import pandas as pd
import glob
import os
from datetime import datetime, timedelta
import psycopg2
import psycopg2.extras
import sqlalchemy
import argparse
import re
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import numpy as np

# Import configuration
sys.path.insert(0, str(Path(__file__).parent))
from config import (
    BHARATI_RAW_DIR, BHARATI_PROCESS_DIR,
    DB_CONNECTION_STRING, DB_CONN_PARAMS,
    ensure_directories_exist, get_logger
)

logger = get_logger(__name__)
ensure_directories_exist()

pd.set_option('future.no_silent_downcasting', True)

# === CONFIGURATION ===
DATA_DIR = str(BHARATI_RAW_DIR)  # Where input files are
TEMPLATE_FILE = str(BHARATI_RAW_DIR / "template.xlsx")
OUTPUT_DIR = str(BHARATI_PROCESS_DIR)  # Where to save output files
process_data_path = str(BHARATI_PROCESS_DIR)

###############################################################################
# Reading each file in a directory and append the data 
###############################################################################

def make_safe_combine(date_obj):
    def safe_combine(time_val):
        try:
            if pd.isna(time_val) or str(time_val).strip() == "":
                return pd.NaT
            time_part = pd.to_datetime(str(time_val), errors='coerce')
            if pd.isna(time_part):
                return pd.NaT
            return datetime.combine(date_obj.date(), time_part.time())
        except Exception:
            return pd.NaT
    return safe_combine

def process_file(file_name):
    print(f"Inserting Data to Database for {file_name}")
    if not os.path.isdir(process_data_path):
        os.mkdir(process_data_path)

    csvname = file_name[0:-5]

    newcsv = open(process_data_path + csvname + '.csv', 'w+')
    dt = datetime.strptime(csvname.strip(), '%d-%b-%Y')
    date = dt.strftime('%m/%d/%Y')
    data = ''
    try:
        csvfile = pd.read_excel(directory_path + file_name , skiprows = range(1, 7))
        for item in csvfile.iterrows():
            data += date+' '+str(item[1]['Unnamed: 2'])[0:5]+','+str(item[1]['Unnamed: 5'])+','+str(item[1]['Unnamed: 6'])+','+str(item[1]['Unnamed: 4'])+','+str(item[1]['Unnamed: 3'])+','+str(item[1]['Unnamed: 7'])+'\n'
    except Exception as e:
        print(e)
    newcsv.write(data)
    newcsv.close()

    # print(file_name)

    """
    item[1]['Unnamed: 2'] = obstime
    item[1]['Unnamed: 5'] = tempr
    item[1]['Unnamed: 6'] = ap
    item[1]['Unnamed: 4'] = ws
    item[1]['Unnamed: 3'] = wd
    item[1]['Unnamed: 7'] = rh

    """
    ###############################################################################
    # Changing nan value into CSV file
    ###############################################################################

    csvfile = csvname + '.csv'
    # for file in file_names:
    file_path = process_data_path + csvfile

    text = open(file_path, "r")
    text = ''.join([i for i in text]).replace("nan", "-999")
    text = ''.join([i for i in text]).replace("--.-", "-999")
    text = ''.join([i for i in text]).replace("--.--", "-999")
    text = ''.join([i for i in text]).replace("----", "-999")    
    text = ''.join([i for i in text]).replace("--", "-999")
    text = ''.join([i for i in text]).replace("-,", "-999,")
    text = ''.join([i for i in text]).replace("-999-999,", "-999,")
    text = ''.join([i for i in text]).replace(" -999,", " 00:00,")    
    x = open(file_path, "w+")
    x.writelines(text)
    x.close()

    ###############################################################################
    # Loading CSV file to Database 
    ###############################################################################
    try:
        conn = psycopg2.connect(**DB_CONN_PARAMS)
        excel_path = os.path.join(DATA_DIR, file_name)
        db_table = 'last_24_hrs_data'
        db_table2 = 'imd_bharati'
        try:
            wb = load_workbook(excel_path, data_only=True)
            ws = wb.active
            b3_value = ws["B3"].value  # e.g., "Date : 05-Jul-2025"

            if b3_value and "Date" in b3_value:
                date_part = b3_value.split(":")[-1].strip()
                date_obj = datetime.strptime(date_part, "%d-%b-%Y")
            else:
                raise ValueError("B3 empty or invalid")

        except Exception as e:
            # Fallback to extracting date from filename
            #print(f"Warning: {e}. Falling back to filename for date.")
            file_stem = os.path.splitext(file_name)[0]
            date_obj = datetime.strptime(file_stem, "%d-%b-%Y")

        # Step 2: Load Excel into DataFrame
        df = pd.read_excel(excel_path, skiprows=5)

        # Rename and filter columns
        column_map = {
            "Time (HH:mm)": "obstime",
            "Temperature (DegC)": "tempr",
            "Presssure (mbar)": "ap",
            "Wind Speed (Knots)": "ws",
            "Wind Direction (Deg)": "wd",
            "Humidity (%)": "rh"
        }
        df = df.rename(columns=column_map)
        df = df[list(column_map.values())]
        # Step 3: Combine date from B3 or fallback with time column
        for col in ['tempr', 'ap', 'ws', 'wd', 'rh']:
            df[col] = df[col].apply(lambda x: np.nan if pd.isna(x) else x)
        numeric_cols = ['tempr', 'ap', 'ws', 'wd', 'rh']
        df[numeric_cols] = df[numeric_cols].apply(pd.to_numeric, errors='coerce')
        columns_to_replace = ['tempr', 'ap', 'ws', 'wd', 'rh']
        for col in columns_to_replace:
            df[col] = df[col].apply(lambda x: x if pd.notna(x) else -999)
        df['obstime'] = df['obstime'].apply(make_safe_combine(date_obj))
        df['obstime'] = pd.to_datetime(df['obstime'], errors='coerce')
        df = df.where(pd.notnull(df), None)

        # Step 4: Insert into DB
        cur = conn.cursor()
        cur.execute(f"""
            SELECT obstime FROM {db_table}
            WHERE obstime >= %s AND obstime < %s
        """, (date_obj, date_obj.replace(hour=23, minute=59, second=59)))
        existing_obstimes = set(row[0] for row in cur.fetchall())

        # Step 5: Split into insert and update DataFrames
        insert_df = df[~df['obstime'].isin(existing_obstimes)].copy()
        update_df = df[df['obstime'].isin(existing_obstimes)].copy()

        # Step 6: Perform bulk INSERT
        if not insert_df.empty:
            data_tuples = [tuple(x) for x in insert_df.to_numpy()]
            insert_query = f"""
                INSERT INTO {db_table} (obstime, tempr, ap, ws, wd, rh)
                VALUES (%s, %s, %s, %s, %s, %s)
            """
            psycopg2.extras.execute_batch(cur, insert_query, data_tuples)

            insert_query2 = insert_query.replace(db_table, db_table2)
            psycopg2.extras.execute_batch(cur, insert_query2, data_tuples)

        # Step 7: Perform bulk UPDATE
        if not update_df.empty:
            should_update = input(f"{len(update_df)} rows already exist. Do you want to update them? (y/n): ").strip().lower()

            while should_update not in ['y', 'n']:
                should_update = input("Please enter 'y' or 'n': ").strip().lower()

            if should_update == 'y':
                update_query = f"""
                    UPDATE {db_table}
                    SET tempr = %s, ap = %s, ws = %s, wd = %s, rh = %s
                    WHERE obstime = %s
                """
                update_tuples = [(
                    row['tempr'], row['ap'], row['ws'], row['wd'], row['rh'], row['obstime']
                ) for _, row in update_df.iterrows()]
                psycopg2.extras.execute_batch(cur, update_query, update_tuples)

                update_query2 = update_query.replace(db_table, db_table2)
                psycopg2.extras.execute_batch(cur, update_query2, update_tuples)

        conn.commit()
        cur.close()

        print(f"Inserted: {len(insert_df)}, Updated: {len(update_df)} from {file_name}")
    except Exception as e:
        print(f"Connection Error {e}")

# === Helper: Extract date from filename ===
def extract_date(filename):
    match = re.search(r'(\d{2}-[A-Za-z]{3}-\d{4})', filename)
    return datetime.strptime(match.group(1), "%d-%b-%Y") if match else None

if __name__ == "__main__":

    parser = argparse.ArgumentParser(description="Process DCWIS data coming from Bharati Station.")
    parser.add_argument("-r","--rerun", action="store_true", help="Rerun for last 5 days data")
    args = parser.parse_args()
    # === Step 1: Gather available files ===
    rw09_files = {extract_date(f): os.path.join(DATA_DIR, f)
                for f in os.listdir(DATA_DIR) if f.startswith("RW09_") and f.endswith(".xlsx")}
    wind_files = {extract_date(f): os.path.join(DATA_DIR, f)
                for f in os.listdir(DATA_DIR) if f.startswith("Wind_") and f.endswith(".xlsx")}

    # === Step 2: Filter common dates and get latest 5 ===
    common_dates = sorted(set(rw09_files) & set(wind_files), reverse=True)[:5]

    # === Step 3: Process each date ===
    for date in common_dates:
        output_filename = date.strftime("%d-%b-%Y") + ".xlsx"
        output_path = os.path.join(OUTPUT_DIR, output_filename)

        # Skip if already exists
        if os.path.exists(output_path) and not args.rerun:
            print(f"Skipping {output_filename} (already exists)")
            continue

        print(f"Processing {output_filename}")

        # === Load input files ===
        rw09_df = pd.read_excel(rw09_files[date])
        wind_df_raw = pd.read_excel(wind_files[date])

        # Clean wind data
        wind_df = wind_df_raw.iloc[1:].copy()
        wind_df.columns = ['Date', 'Time', 'Wind Direction (DEG)', 'WD 2Min', 'WD 10Min',
                        'Wind Speed (knots)', 'WS 2Min', 'WS 10Min',
                        'WD at Max WS in 1 Min (DEG)', 'Max WS in 1Min (knots)']

        # Merge
        merged_df = pd.merge(rw09_df, wind_df, on=['Date', 'Time'], how='inner')

        # Format date
        formatted_date = date.strftime("Date : %d-%b-%Y")

        # Final DataFrame
        final_df = pd.DataFrame({
            "SR.No.": range(1, len(merged_df) + 1),
            "Time (HH:mm)": merged_df["Time"],
            "Wind Direction (Deg)": merged_df["Wind Direction (DEG)"],
            "Wind Speed (Knots)": merged_df["Wind Speed (knots)"],
            "Temperature (DegC)": merged_df["Temperature1MinAvg (DEG C)"],
            "Presssure (mbar)": merged_df["Pressure1MinAvg (mBar)"],
            "Humidity (%)": merged_df["Humidity1MinAvg (%Rh)"],
            "Dew Point (DegC)": merged_df["DewPoint1MinAvg (DEG C)"]
        })

        # Replace missing values
        final_df.fillna({
            "Wind Direction (Deg)": "-",
            "Wind Speed (Knots)": "--.-",
            "Temperature (DegC)": "--.-",
            "Presssure (mbar)": "----.-",
            "Humidity (%)": "--.-",
            "Dew Point (DegC)": "--.-"
        }, inplace=True)

        # Load template and fill in values
        wb = load_workbook(TEMPLATE_FILE)
        ws = wb.active

        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=True), start=6):
            for c_idx, value in enumerate(row, start=2):  # column B = 2
                ws.cell(row=r_idx, column=c_idx, value=value)

        # Fill date into B3
        ws["B3"] = formatted_date

        # Save the file
        wb.save(output_path)
        print(f"Created: {output_filename}")
        process_file(output_filename)
